from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def not_giriniz():
    adsoyad = input('İsminiz: ')
    okulno = input('Okul Numaranız: ')
    sınavpuani1 = input('1. Sınav Notunuz: ')
    sınavpuani2 = input('2. Sınav Notunuz: ')

    with open("bilgiler.txt","a", encoding="utf-8") as file:  # yazılan bilgileri txt ye kayıt ettiriyor   #a a Moduyla dosyaya veri kayıt edicez
        file.write(adsoyad+' '+ okulno+ ' '+sınavpuani1+','+sınavpuani2+'\n') # verileri yazdırıyoruz

def save_al():
    pass
def ortalamalar_globalAI():
    with open("bilgiler.txt","r",encoding="uft-8") as file: #okuma işlemi yapıcaz
        for satir in file:
            print(satir)

while True:
    islem = input('1 - Bilgileri Giriniz\n2 - Notu Kayıt Et\n3 - Notları Görüntüle\n4 - Kapat\n')

    if islem == '1':
        not_giriniz() #not giriş 
    elif islem == '2':
        save_al()     #notları txt ye kayıt ettirecek
    elif islem == '3':
        save_al()
    else:           #yanlış girilen bi numarada fonksiyonu durduracak
        break



data={
    'Joe':{
    'Math':90,
    'History':65,
    'Physics':100,
    'Geography':55,
    },
    'Matilda':{
    'Math':65,
    'History':50,
    'Physics':85,
    'Geography':75,
    },

     'Jon':{
    'Math':100,
    'History':55,
    'Physics':80,
    'Geography':90,
    },
    'Melda':{
    'Math':50,
    'History':95,
    'Physics':90,
    'Geography':55,
    },
    'Ben':{
    'Math':100,
    'History':45,
    'Physics':85,
    'Geography':75,
    }
    
    }


wb=Workbook()
ws=wb.active
ws.title="Grades"

headings=['Name'] + list(data['Joe'].keys())

ws.append(headings)

for person in data:
    grades=list(data[person].values())
    ws.append([person]+grades)
for col in range(2,len(data['Joe'])+2):
    char = get_column_letter(col)
    ws[char+7]=f"SUM({char +'2'}:{char +'6'})/{len(data)}"
for col in range(1,6):
    ws[get_column_letter(col)+1].font=Font(bold=True,color="00008080")        

wb.save('NewGrades.xlsx')   